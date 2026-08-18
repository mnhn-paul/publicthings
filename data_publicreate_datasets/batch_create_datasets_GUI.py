import os
import requests
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
from openpyxl import load_workbook


# ============================================================
# CONFIGURATION
# ============================================================

API_BASE = "https://data.public.lu/api/1"

DATASETS_ENDPOINT = f"{API_BASE}/datasets/"
ME_ENDPOINT = f"{API_BASE}/me/"
ORGANIZATIONS_ENDPOINT = f"{API_BASE}/organizations/"

TIMEOUT = 30


# ============================================================
# DATASET CREATOR
# ============================================================

class DatasetCreatorGUI:

    def __init__(self, root):

        self.root = root

        self.root.title(
            "data.public.lu - Batch Dataset Creator"
        )

        self.root.geometry("850x700")

        # ----------------------------------------------------
        # Variables
        # ----------------------------------------------------

        self.api_key = tk.StringVar()
        self.org_id = tk.StringVar()
        self.excel_file = tk.StringVar()

        # Start with public datasets
        self.is_private = tk.BooleanVar(
            value=False
        )

        # Safety: only create first dataset initially
        self.test_mode = tk.BooleanVar(
            value=True
        )

        # Authenticated user information
        self.current_user = None
        self.current_user_id = None

        self.create_widgets()


    # ========================================================
    # GUI
    # ========================================================

    def create_widgets(self):

        frame = ttk.Frame(
            self.root,
            padding=15
        )

        frame.pack(
            fill="both",
            expand=True
        )


        # ----------------------------------------------------
        # API KEY
        # ----------------------------------------------------

        ttk.Label(
            frame,
            text="API Key:"
        ).grid(
            row=0,
            column=0,
            padx=5,
            pady=5,
            sticky="w"
        )


        ttk.Entry(
            frame,
            textvariable=self.api_key,
            width=60,
            show="*"
        ).grid(
            row=0,
            column=1,
            padx=5,
            pady=5,
            sticky="ew"
        )


        # ----------------------------------------------------
        # ORGANIZATION
        # ----------------------------------------------------

        ttk.Label(
            frame,
            text="Organization ID:"
        ).grid(
            row=1,
            column=0,
            padx=5,
            pady=5,
            sticky="w"
        )


        ttk.Entry(
            frame,
            textvariable=self.org_id,
            width=60
        ).grid(
            row=1,
            column=1,
            padx=5,
            pady=5,
            sticky="ew"
        )


        ttk.Label(
            frame,
            text=(
                "Example: "
                "musee-national-dhistoire-naturelle-luxembourg"
            )
        ).grid(
            row=2,
            column=1,
            padx=5,
            pady=0,
            sticky="w"
        )


        # ----------------------------------------------------
        # PRIVATE
        # ----------------------------------------------------

        ttk.Checkbutton(
            frame,
            text="Create datasets as private/draft",
            variable=self.is_private
        ).grid(
            row=3,
            column=1,
            padx=5,
            pady=10,
            sticky="w"
        )


        # ----------------------------------------------------
        # TEST MODE
        # ----------------------------------------------------

        ttk.Checkbutton(
            frame,
            text=(
                "Test mode - create only the first dataset"
            ),
            variable=self.test_mode
        ).grid(
            row=4,
            column=1,
            padx=5,
            pady=5,
            sticky="w"
        )


        # ----------------------------------------------------
        # EXCEL
        # ----------------------------------------------------

        ttk.Label(
            frame,
            text="Excel File:"
        ).grid(
            row=5,
            column=0,
            padx=5,
            pady=5,
            sticky="w"
        )


        ttk.Entry(
            frame,
            textvariable=self.excel_file,
            width=60
        ).grid(
            row=5,
            column=1,
            padx=5,
            pady=5,
            sticky="ew"
        )


        ttk.Button(
            frame,
            text="Browse",
            command=self.browse_file
        ).grid(
            row=5,
            column=2,
            padx=5,
            pady=5
        )


        # ----------------------------------------------------
        # LOG
        # ----------------------------------------------------

        ttk.Label(
            frame,
            text="Log:"
        ).grid(
            row=6,
            column=0,
            padx=5,
            pady=5,
            sticky="nw"
        )


        self.log_text = tk.Text(
            frame,
            height=25,
            width=100
        )


        self.log_text.grid(
            row=6,
            column=1,
            columnspan=2,
            padx=5,
            pady=5,
            sticky="nsew"
        )


        # ----------------------------------------------------
        # BUTTONS
        # ----------------------------------------------------

        button_frame = ttk.Frame(frame)

        button_frame.grid(
            row=7,
            column=1,
            columnspan=2,
            pady=15
        )


        ttk.Button(
            button_frame,
            text="Test API Connection",
            command=self.test_api
        ).pack(
            side="left",
            padx=5
        )


        ttk.Button(
            button_frame,
            text="Create Dataset(s)",
            command=self.execute
        ).pack(
            side="left",
            padx=5
        )


        # ----------------------------------------------------
        # GRID CONFIGURATION
        # ----------------------------------------------------

        frame.columnconfigure(
            1,
            weight=1
        )

        frame.rowconfigure(
            6,
            weight=1
        )


    # ========================================================
    # LOGGING
    # ========================================================

    def log(self, message):

        print(message)

        self.log_text.insert(
            tk.END,
            message + "\n"
        )

        self.log_text.see(
            tk.END
        )

        self.root.update_idletasks()


    # ========================================================
    # FILE
    # ========================================================

    def browse_file(self):

        filename = filedialog.askopenfilename(
            title="Select Excel File",
            filetypes=[
                ("Excel files", "*.xlsx"),
                ("All files", "*.*")
            ]
        )

        if filename:

            self.excel_file.set(
                filename
            )


    # ========================================================
    # HEADERS
    # ========================================================

    def get_headers(self):

        api_key = self.api_key.get().strip()

        return {
            "X-API-KEY": api_key,
            "Content-Type": "application/json",
            "Accept": "application/json"
        }


    # ========================================================
    # GET CURRENT USER
    # ========================================================

    def get_current_user(self):

        self.log(
            "Getting authenticated user information..."
        )


        try:

            response = requests.get(
                ME_ENDPOINT,
                headers=self.get_headers(),
                timeout=TIMEOUT
            )


        except requests.RequestException as e:

            raise RuntimeError(
                f"Could not connect to data.public.lu: {e}"
            )


        self.log(
            f"GET /me/ -> HTTP {response.status_code}"
        )


        if not response.ok:

            raise RuntimeError(
                "API authentication failed.\n\n"
                f"HTTP {response.status_code}\n\n"
                f"{response.text}"
            )


        try:

            user = response.json()

        except ValueError:

            raise RuntimeError(
                "The /me/ endpoint did not return valid JSON."
            )


        # ----------------------------------------------------
        # Extract user ID
        # ----------------------------------------------------

        user_id = (
            user.get("id")
            or user.get("slug")
        )


        # Sometimes APIs expose the URI instead
        if not user_id:

            uri = user.get("uri")

            if uri:

                user_id = (
                    uri.rstrip("/")
                    .split("/")
                    [-1]
                )


        if not user_id:

            raise RuntimeError(
                "Could not determine the authenticated "
                "user ID from /me/.\n\n"
                f"Response:\n{user}"
            )


        self.current_user = user

        self.current_user_id = user_id


        # ----------------------------------------------------
        # Display user information
        # ----------------------------------------------------

        name = (
            f"{user.get('first_name', '')} "
            f"{user.get('last_name', '')}"
        ).strip()


        if not name:

            name = (
                user.get("email")
                or user.get("slug")
                or str(user_id)
            )


        self.log(
            f"Authenticated user: {name}"
        )

        self.log(
            f"Authenticated user ID: {user_id}"
        )


        return user


    # ========================================================
    # TEST API
    # ========================================================

    def test_api(self):

        api_key = self.api_key.get().strip()


        if not api_key:

            messagebox.showerror(
                "Error",
                "Please enter your API key."
            )

            return


        self.log("")
        self.log("=" * 60)
        self.log("TESTING API")
        self.log("=" * 60)


        try:

            user = self.get_current_user()


            name = (
                f"{user.get('first_name', '')} "
                f"{user.get('last_name', '')}"
            ).strip()


            if not name:

                name = (
                    user.get("email")
                    or str(self.current_user_id)
                )


            messagebox.showinfo(
                "Success",
                "API authentication works.\n\n"
                f"User: {name}\n"
                f"User ID: {self.current_user_id}"
            )


        except Exception as e:

            self.log(
                f"ERROR: {e}"
            )


            messagebox.showerror(
                "API Error",
                str(e)
            )


    # ========================================================
    # CHECK ORGANIZATION
    # ========================================================

    def check_organization(self, org_id):

        self.log(
            f"Checking organization: {org_id}"
        )


        url = (
            f"{ORGANIZATIONS_ENDPOINT}"
            f"{org_id}/"
        )


        try:

            response = requests.get(
                url,
                headers=self.get_headers(),
                timeout=TIMEOUT
            )


        except requests.RequestException as e:

            self.log(
                f"Organization request failed: {e}"
            )

            return None


        self.log(
            f"GET organization -> "
            f"HTTP {response.status_code}"
        )


        if not response.ok:

            self.log(
                response.text
            )

            return None


        try:

            organization = response.json()

        except ValueError:

            self.log(
                "Organization response was not valid JSON."
            )

            return None


        return organization


    # ========================================================
    # CHECK ORGANIZATION MEMBERSHIP
    # ========================================================

    def check_organization_membership(
        self,
        organization
    ):

        if not organization:

            return False


        user_id = self.current_user_id


        if not user_id:

            return False


        self.log(
            "Checking organization membership..."
        )


        members = organization.get(
            "members",
            []
        )


        for member in members:

            user = member.get(
                "user",
                {}
            )


            member_id = (
                user.get("id")
                or user.get("slug")
            )


            if member_id == user_id:

                role = (
                    member.get("role")
                    or member.get("label")
                    or "unknown"
                )


                self.log(
                    f"Organization membership confirmed."
                )

                self.log(
                    f"User role: {role}"
                )

                return True


        # ----------------------------------------------------
        # Some responses may expose a membership flag
        # ----------------------------------------------------

        if organization.get("is_member") is True:

            self.log(
                "Organization membership confirmed."
            )

            return True


        self.log(
            "Could not confirm organization membership "
            "from the organization response."
        )


        return False


    # ========================================================
    # READ EXCEL
    # ========================================================

    def read_excel_file(self, file_path):

        workbook = load_workbook(
            filename=file_path,
            read_only=True,
            data_only=True
        )


        sheet = workbook.active


        rows = list(
            sheet.iter_rows(
                values_only=True
            )
        )


        workbook.close()


        if not rows:

            raise ValueError(
                "The Excel file is empty."
            )


        # ----------------------------------------------------
        # HEADERS
        # ----------------------------------------------------

        headers = []


        for header in rows[0]:

            if header is None:

                headers.append("")

            else:

                headers.append(
                    str(header).strip()
                )


        required_columns = [
            "dataset_title",
            "dataset_description",
            "license",
            "tags"
        ]


        missing = [
            column
            for column in required_columns
            if column not in headers
        ]


        if missing:

            raise ValueError(
                "Missing Excel columns:\n\n"
                + "\n".join(missing)
            )


        # ----------------------------------------------------
        # DATA
        # ----------------------------------------------------

        datasets = []


        for row_number, row in enumerate(
            rows[1:],
            start=2
        ):

            # Ignore completely empty rows
            if all(
                value is None
                or str(value).strip() == ""
                for value in row
            ):

                continue


            dataset = {}


            for i, header in enumerate(headers):

                if not header:

                    continue


                value = (
                    row[i]
                    if i < len(row)
                    else None
                )


                dataset[header] = value


            dataset["_excel_row"] = row_number


            datasets.append(
                dataset
            )


        return datasets


    # ========================================================
    # CLEAN VALUE
    # ========================================================

    def clean_value(self, value):

        if value is None:

            return ""


        return str(value).strip()


    # ========================================================
    # EXTRACT OBJECT ID
    # ========================================================

    def extract_object_id(self, obj):

        if not isinstance(
            obj,
            dict
        ):

            return None


        object_id = (
            obj.get("id")
            or obj.get("slug")
        )


        if object_id:

            return object_id


        uri = obj.get(
            "uri"
        )


        if uri:

            return (
                uri.rstrip("/")
                .split("/")
                [-1]
            )


        return None


    # ========================================================
    # CREATE ONE DATASET
    # ========================================================

    def create_dataset(
        self,
        dataset,
        organization_id,
        private
    ):

        title = self.clean_value(
            dataset.get(
                "dataset_title"
            )
        )


        description = self.clean_value(
            dataset.get(
                "dataset_description"
            )
        )


        license_value = self.clean_value(
            dataset.get(
                "license"
            )
        )


        tags_value = self.clean_value(
            dataset.get(
                "tags"
            )
        )


        if not title:

            raise ValueError(
                "dataset_title is empty."
            )


        if not self.current_user_id:

            raise ValueError(
                "Authenticated user ID is not available."
            )


        # ----------------------------------------------------
        # TAGS
        # ----------------------------------------------------

        if tags_value:

            tags = [
                tag.strip()
                for tag in tags_value.split(",")
                if tag.strip()
            ]

        else:

            tags = []


        # ----------------------------------------------------
        # PAYLOAD
        # ----------------------------------------------------
        #
        # IMPORTANT:
        #
        # owner       = authenticated USER
        #
        # organization = ORGANIZATION
        #
        # These are two different objects.
        #
        # ----------------------------------------------------

        payload = {

            "title": title,

            "description": description,

            "license": license_value,

            "tags": tags,

            "owner": self.current_user_id,

            "organization": organization_id,

            "private": private
        }


        self.log("")
        self.log(
            "Creating dataset:"
        )


        self.log(
            str(payload)
        )


        try:

            response = requests.post(
                DATASETS_ENDPOINT,
                headers=self.get_headers(),
                json=payload,
                timeout=TIMEOUT
            )


        except requests.RequestException as e:

            raise RuntimeError(
                f"Network error: {e}"
            )


        self.log(
            f"POST /datasets/ -> "
            f"HTTP {response.status_code}"
        )


        self.log(
            f"Response: {response.text}"
        )


        # ----------------------------------------------------
        # SUCCESS
        # ----------------------------------------------------

        if response.status_code in (
            200,
            201
        ):

            try:

                result = response.json()

            except ValueError:

                result = {}


            dataset_id = (
                result.get("id")
                or result.get("slug")
                or self.extract_object_id(
                    result
                )
                or "unknown"
            )


            return {
                "success": True,
                "id": dataset_id,
                "response": result
            }


        # ----------------------------------------------------
        # ERROR
        # ----------------------------------------------------

        try:

            error_json = response.json()


            if isinstance(
                error_json,
                dict
            ):

                error_message = (
                    error_json.get("message")
                    or error_json.get("error")
                    or error_json.get("detail")
                    or str(error_json)
                )

            else:

                error_message = str(
                    error_json
                )


        except ValueError:

            error_message = response.text


        raise RuntimeError(
            f"HTTP {response.status_code}: "
            f"{error_message}"
        )


    # ========================================================
    # VERIFY CREATED DATASET
    # ========================================================

    def verify_created_dataset(
        self,
        dataset_id,
        expected_organization_id,
        expected_owner_id
    ):

        if not dataset_id:

            return False


        if dataset_id == "unknown":

            self.log(
                "Cannot verify dataset: "
                "dataset ID is unknown."
            )

            return False


        url = (
            f"{DATASETS_ENDPOINT}"
            f"{dataset_id}/"
        )


        self.log("")
        self.log(
            f"Verifying dataset: {dataset_id}"
        )


        try:

            response = requests.get(
                url,
                headers=self.get_headers(),
                timeout=TIMEOUT
            )


        except requests.RequestException as e:

            self.log(
                f"Verification request failed: {e}"
            )

            return False


        self.log(
            f"GET dataset -> "
            f"HTTP {response.status_code}"
        )


        if not response.ok:

            self.log(
                "Could not retrieve created dataset."
            )

            return False


        try:

            result = response.json()

        except ValueError:

            self.log(
                "Dataset response was not valid JSON."
            )

            return False


        # ----------------------------------------------------
        # Extract owner
        # ----------------------------------------------------

        returned_owner = result.get(
            "owner"
        )


        returned_owner_id = (
            self.extract_object_id(
                returned_owner
            )
        )


        # ----------------------------------------------------
        # Extract organization
        # ----------------------------------------------------

        returned_organization = result.get(
            "organization"
        )


        returned_organization_id = (
            self.extract_object_id(
                returned_organization
            )
        )


        self.log(
            f"Returned owner: "
            f"{returned_owner_id}"
        )


        self.log(
            f"Expected owner: "
            f"{expected_owner_id}"
        )


        self.log(
            f"Returned organization: "
            f"{returned_organization_id}"
        )


        self.log(
            f"Expected organization: "
            f"{expected_organization_id}"
        )


        # ----------------------------------------------------
        # Compare owner
        # ----------------------------------------------------

        owner_ok = (
            returned_owner_id
            == expected_owner_id
        )


        # ----------------------------------------------------
        # Compare organization
        # ----------------------------------------------------

        organization_ok = (
            returned_organization_id
            == expected_organization_id
        )


        if owner_ok:

            self.log(
                "OK: Dataset owner is the "
                "authenticated user."
            )

        else:

            self.log(
                "WARNING: Dataset owner does not "
                "match the authenticated user."
            )


        if organization_ok:

            self.log(
                "OK: Dataset organization is the "
                "selected organization."
            )

        else:

            self.log(
                "WARNING: Dataset organization does "
                "not match the selected organization."
            )


        return (
            owner_ok
            and organization_ok
        )


    # ========================================================
    # EXECUTE
    # ========================================================

    def execute(self):

        api_key = self.api_key.get().strip()

        organization_id = (
            self.org_id.get().strip()
        )

        excel_path = (
            self.excel_file.get().strip()
        )

        private = self.is_private.get()

        test_mode = self.test_mode.get()


        # ----------------------------------------------------
        # BASIC VALIDATION
        # ----------------------------------------------------

        if not api_key:

            messagebox.showerror(
                "Error",
                "API key is required."
            )

            return


        if not organization_id:

            messagebox.showerror(
                "Error",
                "Organization ID is required."
            )

            return


        if not excel_path:

            messagebox.showerror(
                "Error",
                "Excel file is required."
            )

            return


        if not os.path.exists(
            excel_path
        ):

            messagebox.showerror(
                "Error",
                "Excel file does not exist."
            )

            return


        # ----------------------------------------------------
        # START
        # ----------------------------------------------------

        self.log("")
        self.log(
            "=" * 60
        )

        self.log(
            "STARTING"
        )

        self.log(
            "=" * 60
        )


        # ----------------------------------------------------
        # CHECK API + GET USER
        # ----------------------------------------------------

        try:

            user = self.get_current_user()


        except Exception as e:

            messagebox.showerror(
                "API Error",
                str(e)
            )

            return


        # ----------------------------------------------------
        # CHECK ORGANIZATION
        # ----------------------------------------------------

        organization = (
            self.check_organization(
                organization_id
            )
        )


        if organization is None:

            messagebox.showerror(
                "Organization Error",
                "Could not verify organization:\n\n"
                f"{organization_id}\n\n"
                "The dataset creation has been stopped."
            )

            return


        organization_name = (
            organization.get("name")
            or organization_id
        )


        self.log(
            f"Organization found: "
            f"{organization_name}"
        )


        # ----------------------------------------------------
        # MEMBERSHIP
        # ----------------------------------------------------

        membership_confirmed = (
            self.check_organization_membership(
                organization
            )
        )


        if not membership_confirmed:

            answer = messagebox.askyesno(
                "Membership not confirmed",
                "The API response did not allow the "
                "program to confirm that your user is "
                "a member of this organization.\n\n"
                f"Organization:\n{organization_name}\n\n"
                f"User ID:\n{self.current_user_id}\n\n"
                "The data.public.lu API requires appropriate "
                "organization permissions for write operations.\n\n"
                "Do you want to continue anyway?"
            )


            if not answer:

                return


        # ----------------------------------------------------
        # READ EXCEL
        # ----------------------------------------------------

        try:

            datasets = self.read_excel_file(
                excel_path
            )


        except Exception as e:

            messagebox.showerror(
                "Excel Error",
                str(e)
            )

            return


        if not datasets:

            messagebox.showerror(
                "Excel Error",
                "No datasets found in the Excel file."
            )

            return


        self.log(
            f"Found {len(datasets)} dataset(s)."
        )


        # ----------------------------------------------------
        # TEST MODE
        # ----------------------------------------------------

        if test_mode:

            datasets_to_create = (
                datasets[:1]
            )


            self.log(
                "TEST MODE ENABLED."
            )


            self.log(
                "Only the first dataset will be created."
            )


        else:

            datasets_to_create = (
                datasets
            )


            confirmation = messagebox.askyesno(
                "Confirm batch creation",
                f"You are about to create "
                f"{len(datasets_to_create)} datasets.\n\n"
                f"Organization:\n"
                f"{organization_name}\n\n"
                f"Organization ID:\n"
                f"{organization_id}\n\n"
                f"Owner / User ID:\n"
                f"{self.current_user_id}\n\n"
                f"Private:\n"
                f"{private}\n\n"
                f"Continue?"
            )


            if not confirmation:

                return


        # ----------------------------------------------------
        # CREATE DATASETS
        # ----------------------------------------------------

        successful = []

        failed = []


        for index, dataset in enumerate(
            datasets_to_create,
            start=1
        ):

            title = self.clean_value(
                dataset.get(
                    "dataset_title"
                )
            )


            excel_row = dataset.get(
                "_excel_row",
                "?"
            )


            self.log("")


            self.log(
                f"[{index}/"
                f"{len(datasets_to_create)}]"
                f" Excel row {excel_row}: "
                f"{title}"
            )


            try:

                result = self.create_dataset(
                    dataset,
                    organization_id,
                    private
                )


                dataset_id = (
                    result["id"]
                )


                # ------------------------------------------------
                # Verify owner and organization
                # ------------------------------------------------

                verification_ok = (
                    self.verify_created_dataset(
                        dataset_id,
                        organization_id,
                        self.current_user_id
                    )
                )


                successful.append(
                    {
                        "title": title,
                        "id": dataset_id,
                        "verified": verification_ok
                    }
                )


                if verification_ok:

                    self.log(
                        f"SUCCESS + VERIFIED: "
                        f"{dataset_id}"
                    )

                else:

                    self.log(
                        f"SUCCESS but verification "
                        f"needs attention: "
                        f"{dataset_id}"
                    )


            except Exception as e:

                failed.append(
                    {
                        "title": title,
                        "error": str(e)
                    }
                )


                self.log(
                    f"FAILED: {e}"
                )


        # ----------------------------------------------------
        # FINAL REPORT
        # ----------------------------------------------------

        self.log("")

        self.log(
            "=" * 60
        )

        self.log(
            "FINISHED"
        )

        self.log(
            "=" * 60
        )


        self.log(
            f"Successful: "
            f"{len(successful)}"
        )


        self.log(
            f"Failed: "
            f"{len(failed)}"
        )


        # ----------------------------------------------------
        # SUCCESSFUL DATASETS
        # ----------------------------------------------------

        if successful:

            self.log("")
            self.log(
                "SUCCESSFUL DATASETS:"
            )


            for item in successful:

                verification = (
                    "VERIFIED"
                    if item["verified"]
                    else "NOT VERIFIED"
                )


                self.log(
                    f"- {item['title']} "
                    f"-> {item['id']} "
                    f"({verification})"
                )


        # ----------------------------------------------------
        # FAILED DATASETS
        # ----------------------------------------------------

        if failed:

            self.log("")
            self.log(
                "FAILED DATASETS:"
            )


            for item in failed:

                self.log(
                    f"- {item['title']}: "
                    f"{item['error']}"
                )


        # ----------------------------------------------------
        # SUMMARY
        # ----------------------------------------------------

        summary = (
            f"Finished.\n\n"
            f"Successful: {len(successful)}\n"
            f"Failed: {len(failed)}\n\n"
            f"Organization:\n"
            f"{organization_name}\n\n"
            f"Organization ID:\n"
            f"{organization_id}\n\n"
            f"Owner / User ID:\n"
            f"{self.current_user_id}"
        )


        if test_mode:

            summary += (
                "\n\nTEST MODE was enabled."
                "\nOnly the first Excel row was processed."
            )


        messagebox.showinfo(
            "Finished",
            summary
        )


# ============================================================
# MAIN
# ============================================================

if __name__ == "__main__":

    root = tk.Tk()

    app = DatasetCreatorGUI(
        root
    )

    root.mainloop()
